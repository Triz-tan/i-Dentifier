import customtkinter as ctk  # Import the customtkinter library as ctk
import style_register as style  # Import the style_register module as style
import cv2  # Import the OpenCV library for computer vision tasks
from PIL import Image, ImageTk  # Import the PIL (Python Imaging Library) for image processing
import os  # Import the os module for interacting with the operating system
from tkinter import messagebox  # Import the messagebox module from tkinter for displaying messages
import openpyxl  # Import the openpyxl library for working with Excel files


# Define a class named 'register_window' that inherits from 'ctk.CTkToplevel'
class register_window(ctk.CTkToplevel):
    
    def __init__(self, master):
        super().__init__(master)
        self.geometry("1366x768")
        self.title("Register")
        self.attributes('-topmost', True)
        self.attributes('-fullscreen', True)

        # Load the image file
        self.image = Image.open("GUI_images/RegisterBackground.png")

        # Create a PhotoImage object from the image
        self.photo = ImageTk.PhotoImage(self.image)

        # Create a canvas widget with the background image
        self.canvas = ctk.CTkCanvas(self, width=self.image.width, height=self.image.height)
        self.canvas.create_image(0, 0, anchor=ctk.NW, image=self.photo)
        self.canvas.pack()
        
        # Create back button with style
        self.back_btn = style.back_btn_style(self, 'Back', self.back_func_btn)
        self.back_btn.place(x=1122.5, y=652.4)

        # Create entry fields with style
        self.name_entry = style.entry_style(self, "Enter your full name")
        self.name_entry.place(x=836.9, y=228)

        self.user_entry = style.entry_style(self, "Enter your user Id (Use number only)")
        self.user_entry.place(x=836.9, y=310.6)

        self.position_entry = style.entry_style(self, "Enter your position")
        self.position_entry.place(x=836.9, y=397.7)

        self.status_entry = style.entry_style(self, "Enter your status")
        self.status_entry.place(x=836.9, y=480.3)

        # Create capture button with style
        self.capture_btn = style.register_btn_style(self, "Capture", self.capture_func_btn)
        self.capture_btn.place(x=954.5, y=546.9)

        # Create webcam label with style
        self.webcam_label = style.webcam_style_label(self)
        self.webcam_label.place(x=109.4, y=193, width=598.8, height=428.7)
        
        # Initialize variables
        self.name = ""
        self.user_id = ""
        self.position = ""
        self.status = ""
        self.capture_image_path = None
        self.btn_confirm = False

        # Add webcam to the webcam label
        self.add_webcam(self.webcam_label)
    
    # Method to add webcam to the label
    def add_webcam(self, label):
        if 'capture' not in self.__dict__:
            self.capture = cv2.VideoCapture(0)

        self._label = label
        self.process_webcam()

    # Method to process the webcam frames
    def process_webcam(self):
        ret, frame = self.capture.read()
        self.most_recent_capture_arr = frame
        img_ = cv2.cvtColor(self.most_recent_capture_arr, cv2.COLOR_BGR2RGB)
        self.most_recent_capture_pil = Image.fromarray(img_)
        imgtk = ImageTk.PhotoImage(image=self.most_recent_capture_pil)
        self._label.imgtk = imgtk
        self._label.configure(image=imgtk)
        self._label.after(20, self.process_webcam)
    
    # Method to capture the image from the webcam
    def capture_func_btn(self):
        print("Capture function button just clicked")
        
        user_id_value = self.user_entry.get() 

        if not user_id_value.isnumeric():
            self.attributes('-topmost', False)
            messagebox.showerror('Error', "Please use numbers only for your ID")
            self.attributes('-topmost', True)
            return

        ret, frame = self.capture.read()
        if ret:
            if not os.path.exists('Registerimage'):
                os.makedirs('Registerimage')
            path_image = 'Registerimage/{}.jpg'.format(user_id_value) 

            if os.path.exists(path_image):
                self.attributes('-topmost', False)
                messagebox.showerror('Error', "The ID is already used")
                self.attributes('-topmost', True)
                return

            cv2.imwrite(path_image, frame)

            # Create a new window for image registration
            self.photo_register = ctk.CTkToplevel()
            self.photo_register.geometry("1366x768")
            self.photo_register.title("Photo Registration")
            self.photo_register.attributes('-topmost', True)
            self.photo_register.attributes('-fullscreen', True) 
            self.attributes('-topmost', False)
                
            # Load the image file
            self.image_confirm = Image.open("GUI_images/RegisterBackgroundConfirm.png")

            # Create a PhotoImage object from the image
            self.photo_confirm = ImageTk.PhotoImage(self.image_confirm)

            # Create a canvas widget with the background image
            self.canvas_confirm = ctk.CTkCanvas(self.photo_register, width=self.image_confirm.width, height=self.image_confirm.height)
            self.canvas_confirm.create_image(0, 0, anchor=ctk.NW, image=self.photo_confirm)
            self.canvas_confirm.pack()

            # Create webcam label for confirmation
            self.webcam_label = style.webcam_style_label2(self.photo_register)
            self.webcam_label.place(x=663.4, y=197.4, width=598.8, height=428.7)

            # Store capture image path
            self.capture_image_path = path_image  
            self.captured_image = Image.open(path_image)
            self.captured_image = ImageTk.PhotoImage(self.captured_image)
            self.webcam_label.configure(image=self.captured_image)
            self.webcam_label.image = self.captured_image 

            # Create 'Yes' and 'No' buttons for confirmation
            self.yes_btn = style.register_btn_style(self.photo_register, "Yes", self.yes_func_btn)
            self.yes_btn.place(x=178.3, y=526.4)

            self.no_btn = style.regno_btn_style(self.photo_register, "No", self.no_func_btn)
            self.no_btn.place(x=348.8, y=526.4)
            
    # Method to handle 'Yes' button click
    def yes_func_btn(self):
        print("Yes button clicked")
                
        self.name = self.name_entry.get()
        self.user_id = self.user_entry.get()
        self.position = self.position_entry.get()
        self.status = self.status_entry.get()

        workbook = openpyxl.load_workbook("Csc_Form_48.xlsx")
        worksheet = workbook.get_sheet_by_name("Personnel")
        
        row = worksheet.max_row + 1

        worksheet.cell(row=row, column=1, value=self.name)
        worksheet.cell(row=row, column=2, value=self.user_id)
        worksheet.cell(row=row, column=3, value=self.position)
        worksheet.cell(row=row, column=4, value=self.status)
        workbook.save("Csc_Form_48.xlsx")

        self.photo_register.destroy()
        self.capture.release()
        
        messagebox.showinfo("Register Entry", "You are now officially registered.")

        # Method to handle 'No' button click
    def no_func_btn(self):
        self.btn_confirm = False
        os.remove(self.capture_image_path)
        self.reset_capture()
        
    # Method to reset the capture
    def reset_capture(self):
        self.webcam_label.configure(image=None)
        self.webcam_label.image = None
        self.photo_register.destroy()
        
    # Method to handle 'Back' button click
    def back_func_btn(self):
        self.capture.release()
        self.destroy()

