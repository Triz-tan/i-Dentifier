# The above code imports various libraries and modules required for the program.
import customtkinter as ctk
import style_register as style
import cv2
from PIL import Image, ImageTk
import os
from tkinter import messagebox
from datetime import datetime
import openpyxl
import face_recognition
import numpy as np
from tensorflow.keras.models import model_from_json
from tensorflow.keras.preprocessing.image import img_to_array

os.environ['CUDA_VISIBLE_DEVICES'] = ''

# The above code defines the `face_recog_window` class and initializes it with certain attributes, such as window size, title, and fullscreen mode.
# It also sets the path to the folder containing registered employee images and reads each image from that folder, storing them in a list along with the respective employee IDs.
# Additionally, it displays an information message box indicating that the system is generating images for the registered employees.
class face_recog_window(ctk.CTkToplevel):
    def __init__(self, master):
        super().__init__(master)
        self.geometry("1366x768")
        self.title("Face Recognition")
        self.attributes('-fullscreen', True)
        self.attributes('-topmost', False)

        self.path_image = 'Registerimage'
        images = []
        self.employee_id = []
        self.mylist = os.listdir(self.path_image)

        messagebox.showinfo("Generating Registered Images", "The system is generating images for the registered employees. Please wait for a moment. Thank you.")
        for cl in self.mylist:
            curImg = cv2.imread(f'{self.path_image}/{cl}')
            images.append(curImg)
            self.employee_id.append(os.path.splitext(cl)[0])

        # Call the `findeEncodings` function with the `images` list as input
        # and assign the returned list of face encodings to `self.encodeListKnown`
        def findeEncodings(images):
            encodeList = []
            for img in images:
                img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                encode = face_recognition.face_encodings(img)[0]
                encodeList.append(encode)
            return encodeList

        self.encodeListKnown = findeEncodings(images)
     

        self.attributes('-topmost', True)
        # Load the image file
        self.image_face = Image.open("GUI_images/AttendanceBackground.png")

        # Create a PhotoImage object from the image
        self.photo_face = ImageTk.PhotoImage(self.image_face)

        # Create a canvas widget with the background image
        self.canvas_face = ctk.CTkCanvas(self, width=self.image_face.width, height=self.image_face.height)
        self.canvas_face.create_image(0, 0, anchor=ctk.NW, image=self.photo_face)
        self.canvas_face.pack()

        
        self.back_btn = style.back_btn_style(self, 'Back', self.back_func_btn)
        self.back_btn.place(x=1122.5, y=652.4)

        self.arrival_btn_AM = style.frame_btn_style(self, 'Arrival AM', self.arrival_func_AM)
        self.arrival_btn_AM.place(x=864.3, y=385.3)
        
        self.departure_btn_AM = style.frame_btn_style(self, 'Departure AM', self.departure_func_AM)
        self.departure_btn_AM.place(x=1046.6, y=385.3)
        
        self.arrival_btn_PM = style.frame_btn_style(self, 'Arrival PM', self.arrival_func_PM)
        self.arrival_btn_PM.place(x=864.3, y=523.6)

        self.departure_btn_PM = style.frame_btn_style(self, 'Departure PM', self.departure_func_PM)
        self.departure_btn_PM.place(x=1046.6, y=523.6)
        
        

        self.webcam_label = style.webcam_style_label3(self)
        self.webcam_label.place(x=109.4, y=193, width=598.8, height=428.7)


        self.add_webcam(self.webcam_label)
    # The add_webcam method is responsible for capturing video from a webcam and updating the specified label with the captured frames.
    def add_webcam(self, label):
        if 'capture' not in self.__dict__:
            self.capture = cv2.VideoCapture(0)

        self._label = label
        self.process_webcam()
    # The process_webcam method processes the captured frame from the webcam and updates the label with the latest frame.
    # It converts the frame from BGR to RGB color space, creates a PIL Image object, converts it to ImageTk format, and assigns it to the label's image attribute.
    # Finally, it uses the 'after' method to schedule the 'process_webcam' method to run again after 20 milliseconds.
    def process_webcam(self):
        ret, frame = self.capture.read()

        self.most_recent_capture_arr = frame
        self.img_ = cv2.cvtColor(self.most_recent_capture_arr, cv2.COLOR_BGR2RGB)
        self.most_recent_capture_pil = Image.fromarray(self.img_)
        imgtk = ImageTk.PhotoImage(image=self.most_recent_capture_pil)
        self._label.imgtk = imgtk
        self._label.configure(image=imgtk)
        self._label.after(20, self.process_webcam)
           

    def arrival_func_AM(self):
        # Load Face Detection Model
        face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
        # Load Anti-Spoofing Model graph
        json_file = open('antispoofing_model.json','r')
        loaded_model_json = json_file.read()
        json_file.close()
        model = model_from_json(loaded_model_json)
        # load antispoofing model weights 
        model.load_weights('antispoofing_model.h5')
        
        #If spoofing is detected, the method returns and further processing is stopped. 
        faces = face_cascade.detectMultiScale(self.img_ ,1.3,5)
        for (x,y,w,h) in faces:  
            face = self.most_recent_capture_arr
            resized_face = cv2.resize(face,(160,160))
            resized_face = resized_face.astype("float") / 255.0
            resized_face = np.expand_dims(resized_face, axis=0)
            preds = model.predict(resized_face)[0]
            if preds> 0.05:
                self.attributes('-topmost', False)
                messagebox.showerror("Error", "You are spoofing.")
                self.attributes('-topmost', True)
                return
          
 
        facesCurFrame = face_recognition.face_locations(self.img_)
        encodesCurFrame = face_recognition.face_encodings(self.img_, facesCurFrame) 

        for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
            # Compare the face encoding with the known face encodings
            matches = face_recognition.compare_faces(self.encodeListKnown, encodeFace)
            # Calculate the face distance between the face encoding and the known face encodings
            faceDis = face_recognition.face_distance(self.encodeListKnown, encodeFace)
            # Find the index of the smallest face distance
            matchIndex = np.argmin(faceDis)

             # If the matchIndex is within the valid range of the known face encodings and there is a match
            if matchIndex < len(self.encodeListKnown) and matches[matchIndex]:
                 # Get the recognized employee's ID (converted to uppercase)
                self.facerecognized = self.employee_id[matchIndex].upper()
                        
                image_path = None
                # Search for the corresponding image file path based on the recognized employee's ID
                for path in self.mylist:
                    if os.path.splitext(path)[0] == self.facerecognized:
                        # Create the complete image path by joining the path_image directory with the filename
                        image_path = os.path.join(self.path_image, path)
                        break

                if image_path is not None:
                    image = Image.open(image_path)
                    photo = ImageTk.PhotoImage(image)

                    self.arrival_am = ctk.CTkToplevel()
                    self.arrival_am.geometry("1366x768")
                    self.arrival_am.title("Arrival PM")
                    self.arrival_am.attributes('-topmost', True)
                    self.arrival_am.attributes('-fullscreen', True)
                    self.attributes('-topmost', False)

                    # Load the image file
                    self.image_am = Image.open("GUI_images/AttendanceBackgroundConfirm.png")

                    # Create a PhotoImage object from the image
                    self.photo_am = ImageTk.PhotoImage(self.image_am)

                    # Create a canvas widget with the background image
                    self.canvas_am = ctk.CTkCanvas(self.arrival_am, width=self.image_am.width,
                                                height=self.image_am.height)
                    self.canvas_am.create_image(0, 0, anchor=ctk.NW, image=self.photo_am)
                    self.canvas_am.pack()

                    self.webcam_label2 = style.webcam_style_label4(self.arrival_am)
                    self.webcam_label2.place(x=663.4, y=197.4, width=598.8, height=428.7)

                    self.webcam_label2.configure(image=photo)
                    self.webcam_label2.image = photo

                    self.confirm_btn_am = style.frame_btn_style(self.arrival_am, "Confirm",
                                                                self.confirm_func_btn_am)
                    self.confirm_btn_am.place(x=183.6, y=516.2)

                    self.reset_btn_am = style.no_btn_style(self.arrival_am, "Reset", self.reset_func_btn_am)
                    self.reset_btn_am.place(x=354.1, y=516.2)
                else:
                    self.attributes('-topmost', False)
                    messagebox.showerror("Error", "No matching image found.")
                    self.attributes('-topmost', True)
            else:
                self.attributes('-topmost', False)
                messagebox.showerror("Error", "Unknown person detected.")
                self.attributes('-topmost', True)

                
                                  

    def confirm_func_btn_am(self):

        workbook = openpyxl.load_workbook("Csc_Form_48.xlsx")
        worksheet1 = workbook.get_sheet_by_name("ArvAM")
        worksheet2 = workbook.get_sheet_by_name("ID,Day,Date")
                    
        now = datetime.now()
        row = worksheet1.max_row + 1
        row2 = worksheet2.max_row + 1
                    
        worksheet1.cell(row=row, column=4, value=now.strftime("%H:%M:%S"))
        worksheet1.cell(row=row, column=1, value=self.facerecognized)
        worksheet1.cell(row=row, column=2, value=now.strftime("%m/%d/%Y"))
        worksheet1.cell(row=row, column=3, value=now.strftime("%a"))


        worksheet2.cell(row=row2, column=1, value=self.facerecognized)
        worksheet2.cell(row=row2, column=3, value=now.strftime("%m/%d/%Y"))
        worksheet2.cell(row=row2, column=4, value=now.strftime("%a"))

        workbook.save('Csc_Form_48.xlsx')
        self.arrival_am.attributes('-topmost', False)
        messagebox.showinfo("Arrival", "You arrived at " + now.strftime("%A") + " , " + now.strftime("%H:%M"))
        self.attributes('-topmost', True)
        self.arrival_am.destroy()
        
    
    def reset_func_btn_am(self):
        self.btn_confirmed = False
        self.webcam_label2.configure(image=None)
        self.webcam_label2.image = None
        self.arrival_am.destroy()

    
        print("Reset just clicked")
        

   
    def arrival_func_PM(self):  
        # Load Face Detection Model
        face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
        # Load Anti-Spoofing Model graph
        json_file = open('antispoofing_model.json','r')
        loaded_model_json = json_file.read()
        json_file.close()
        model = model_from_json(loaded_model_json)
        # load antispoofing model weights 
        model.load_weights('antispoofing_model.h5')
        print("Model loaded from disk")
        
        faces = face_cascade.detectMultiScale(self.img_ ,1.3,5)
        for (x,y,w,h) in faces:  
            face = self.most_recent_capture_arr
            resized_face = cv2.resize(face,(160,160))
            resized_face = resized_face.astype("float") / 255.0
            resized_face = np.expand_dims(resized_face, axis=0)
            preds = model.predict(resized_face)[0]
            if preds> 0.01:
                self.attributes('-topmost', False)
                messagebox.showerror("Error", "You are spoofing.")
                self.attributes('-topmost', True)
                return        
       
        facesCurFrame = face_recognition.face_locations(self.img_)
        encodesCurFrame = face_recognition.face_encodings(self.img_, facesCurFrame)

        if len(encodesCurFrame) > 1:
            self.attributes('-topmost', False)
            messagebox.showerror("Error", "Please only one person at a time.")
            self.attributes('-topmost', True)
            return
        
        for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
            matches = face_recognition.compare_faces(self.encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(self.encodeListKnown, encodeFace)

            matchIndex = np.argmin(faceDis)

            if matchIndex < len(self.encodeListKnown) and matches[matchIndex]:
                self.facerecognized = self.employee_id[matchIndex].upper()
                        
                
                image_path = None
                for path in self.mylist:
                    if os.path.splitext(path)[0] == self.facerecognized:
                        image_path = os.path.join(self.path_image, path)
                        break

                if image_path is not None:
                    image = Image.open(image_path)
                    photo = ImageTk.PhotoImage(image)
                    
                    self.arrival_pm = ctk.CTkToplevel()
                    self.arrival_pm.geometry("1366x768")
                    self.arrival_pm.title("Arrival PM")
                    self.arrival_pm.attributes('-topmost', True)
                    self.arrival_pm.attributes('-fullscreen', True) 
                    self.attributes('-topmost', False)

                    # Load the image file
                    self.image_pm = Image.open("GUI_images/AttendanceBackgroundConfirm.png")

                    # Create a PhotoImage object from the image
                    self.photo_pm = ImageTk.PhotoImage(self.image_pm)

                    # Create a canvas widget with the background image
                    self.canvas_pm = ctk.CTkCanvas(self.arrival_pm, width=self.image_pm.width, height=self.image_pm.height)
                    self.canvas_pm.create_image(0, 0, anchor=ctk.NW, image=self.photo_pm)
                    self.canvas_pm.pack()

                    self.webcam_label2 = style.webcam_style_label4(self.arrival_pm)
                    self.webcam_label2.place(x=663.4, y=197.4, width=598.8, height=428.7)

                    self.webcam_label2.configure(image=photo)
                    self.webcam_label2.image = photo

                    self.confirm_btn_pm = style.frame_btn_style(self.arrival_pm, "Confirm", self.confirm_func_btn_pm)
                    self.confirm_btn_pm.place(x=183.6, y=516.2)

                    self.reset_btn_pm = style.no_btn_style(self.arrival_pm, "Reset", self.reset_func_btn_pm)
                    self.reset_btn_pm.place(x=354.1, y=516.2)
                else:
                    self.attributes('-topmost', False)
                    messagebox.showerror("Error", "No matching image found.")
                    self.attributes('-topmost', True)
                    
    
            else:
                self.attributes('-topmost', False)
                messagebox.showerror("Error", "Unknown Face Detected.")
                self.attributes('-topmost', True)

    def confirm_func_btn_pm(self):


        workbook = openpyxl.load_workbook("Csc_Form_48.xlsx")
        worksheet1 = workbook.get_sheet_by_name("ArvPM")
        worksheet2 = workbook.get_sheet_by_name("ID,Day,Date")
        now = datetime.now()
        row = worksheet1.max_row + 1
        row2 = worksheet2.max_row + 1

        worksheet1.cell(row=row, column=4, value=now.strftime("%H:%M:%S"))
        worksheet1.cell(row=row, column=1, value=self.facerecognized)
        worksheet1.cell(row=row, column=2, value=now.strftime("%m/%d/%Y"))
        worksheet1.cell(row=row, column=3, value=now.strftime("%a"))

                   # Check if the employee ID already exists in the "ID,Day,Date" sheet for the current day
        id_exists = False
        for row in worksheet2.iter_rows(min_row=2, max_col=4, values_only=True):
            if row[0] == self.facerecognized and row[2] == now.strftime("%m/%d/%Y"):
                id_exists = True
                break

        if not id_exists:
            worksheet2.cell(row=row2, column=2, value=self.facerecognized)
            worksheet2.cell(row=row2, column=3, value=now.strftime("%m/%d/%Y"))
            worksheet2.cell(row=row2, column=4, value=now.strftime("%a"))

        workbook.save('Csc_Form_48.xlsx')
        self.arrival_pm.attributes('-topmost', False)
        messagebox.showinfo("Arrival", "You arrived at " + now.strftime("%A") + " , " + now.strftime("%H:%M"))
        self.attributes('-topmost', True)
        self.arrival_pm.destroy()
        
        
    
    def reset_func_btn_pm(self):
        self.btn_confirmed = False
        self.webcam_label2.configure(image=None)
        self.webcam_label2.image = None
        self.arrival_pm.destroy()                                



    def departure_func_AM(self):
        # Load Face Detection Model
        face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
        # Load Anti-Spoofing Model graph
        json_file = open('antispoofing_model.json','r')
        loaded_model_json = json_file.read()
        json_file.close()
        model = model_from_json(loaded_model_json)
        # load antispoofing model weights 
        model.load_weights('antispoofing_model.h5')
        print("Model loaded from disk")
        
        faces = face_cascade.detectMultiScale(self.img_ ,1.3,5)
        for (x,y,w,h) in faces:  
            face = self.most_recent_capture_arr
            resized_face = cv2.resize(face,(160,160))
            resized_face = resized_face.astype("float") / 255.0
            resized_face = np.expand_dims(resized_face, axis=0)
            preds = model.predict(resized_face)[0]
            if preds> 0.01:
                self.attributes('-topmost', False)
                messagebox.showerror("Error", "You are spoofing.")
                self.attributes('-topmost', True)
                return
            
        facesCurFrame = face_recognition.face_locations(self.img_)
        encodesCurFrame = face_recognition.face_encodings(self.img_, facesCurFrame)

        if len(encodesCurFrame) > 1:
            self.attributes('-topmost', False)
            messagebox.showerror("Error", "Please only one person at a time.")
            self.attributes('-topmost', True)
            return

        for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
            matches = face_recognition.compare_faces(self.encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(self.encodeListKnown, encodeFace)

            matchIndex = np.argmin(faceDis)

            if matchIndex < len(self.encodeListKnown) and matches[matchIndex]:
                self.facerecognized = self.employee_id[matchIndex].upper()
                
                image_path = None
                for path in self.mylist:
                    if os.path.splitext(path)[0] == self.facerecognized:
                        image_path = os.path.join(self.path_image, path)
                        break

                if image_path is not None:
                    image = Image.open(image_path)
                    photo = ImageTk.PhotoImage(image)
                    
                    self.dep_am = ctk.CTkToplevel()
                    self.dep_am.geometry("1366x768")
                    self.dep_am.title("Departure AM")
                    self.dep_am.attributes('-topmost', True)
                    self.dep_am.attributes('-fullscreen', True) 
                    self.attributes('-topmost', False)

                    # Load the image file
                    self.image_depam = Image.open("GUI_images/AttendanceBackgroundConfirm.png")

                    # Create a PhotoImage object from the image
                    self.photo_depam = ImageTk.PhotoImage(self.image_depam)

                    # Create a canvas widget with the background image
                    self.canvas_depam = ctk.CTkCanvas(self.dep_am, width=self.image_depam.width, height=self.image_depam.height)
                    self.canvas_depam.create_image(0, 0, anchor=ctk.NW, image=self.photo_depam)
                    self.canvas_depam.pack()

                    self.webcam_label2 = style.webcam_style_label4(self.dep_am)
                    self.webcam_label2.place(x=663.4, y=197.4, width=598.8, height=428.7)

                    self.webcam_label2.configure(image=photo)
                    self.webcam_label2.image = photo

                    self.confirm_btn_depam = style.frame_btn_style(self.dep_am, "Confirm", self.confirm_func_btn_depam)
                    self.confirm_btn_depam.place(x=183.6, y=516.2)

                    self.reset_btn_pm = style.no_btn_style(self.dep_am, "Reset", self.reset_func_btn_depam)
                    self.reset_btn_pm.place(x=354.1, y=516.2)
                else:
                    self.attributes('-topmost', False)
                    messagebox.showerror("Error", "No matching image found.")
                    self.attributes('-topmost', True)
                    
    
            else:
                self.attributes('-topmost', False)
                messagebox.showerror("Error", "Unknown Face Detected.")
                self.attributes('-topmost', True)

    def confirm_func_btn_depam(self):
        
        workbook = openpyxl.load_workbook("Csc_Form_48.xlsx")
        worksheet = workbook.get_sheet_by_name("DepAM")
        now = datetime.now()
        row = worksheet.max_row + 1
                    
        worksheet.cell(row=row, column=4, value=now.strftime("%H:%M:%S"))
        worksheet.cell(row=row, column=1, value=self.facerecognized)
        worksheet.cell(row=row, column=2, value=now.strftime("%m/%d/%Y"))
        worksheet.cell(row=row, column=3, value=now.strftime("%a"))

        workbook.save('Csc_Form_48.xlsx')
        self.dep_am.attributes('-topmost', False)
        messagebox.showinfo("Arrival", "You departed at " + now.strftime("%A") + " , " + now.strftime("%H:%M"))
        self.attributes('-topmost', True)
        self.dep_am.destroy()
        
       
    def reset_func_btn_depam(self):
        self.btn_confirmed = False
        self.webcam_label2.configure(image=None)
        self.webcam_label2.image = None
        self.dep_am.destroy()

    def departure_func_PM(self):
        # Load Face Detection Model
        face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
        # Load Anti-Spoofing Model graph
        json_file = open('antispoofing_model.json','r')
        loaded_model_json = json_file.read()
        json_file.close()
        model = model_from_json(loaded_model_json)
        # load antispoofing model weights 
        model.load_weights('antispoofing_model.h5')
        print("Model loaded from disk")
        
        faces = face_cascade.detectMultiScale(self.img_ ,1.3,5)
        for (x,y,w,h) in faces:  
            face = self.most_recent_capture_arr
            resized_face = cv2.resize(face,(160,160))
            resized_face = resized_face.astype("float") / 255.0
            resized_face = np.expand_dims(resized_face, axis=0)
            preds = model.predict(resized_face)[0]
            if preds> 0.01:
                self.attributes('-topmost', False)
                messagebox.showerror("Error", "You are spoofing.")
                self.attributes('-topmost', True)
                return
            
        facesCurFrame = face_recognition.face_locations(self.img_)
        encodesCurFrame = face_recognition.face_encodings(self.img_, facesCurFrame)

        if len(encodesCurFrame) > 1:
            self.attributes('-topmost', False)
            messagebox.showerror("Error", "Please only one person at a time.")
            self.attributes('-topmost', True)
            return
        
        for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
            matches = face_recognition.compare_faces(self.encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(self.encodeListKnown, encodeFace)

            matchIndex = np.argmin(faceDis)

            if matchIndex < len(self.encodeListKnown) and matches[matchIndex]:
                self.facerecognized = self.employee_id[matchIndex].upper()
                
                image_path = None
                for path in self.mylist:
                    if os.path.splitext(path)[0] == self.facerecognized:
                        image_path = os.path.join(self.path_image, path)
                        break

                if image_path is not None:
                    image = Image.open(image_path)
                    photo = ImageTk.PhotoImage(image)
                    
                    self.dep_pm = ctk.CTkToplevel()
                    self.dep_pm.geometry("1366x768")
                    self.dep_pm.title("Departure AM")
                    self.dep_pm.attributes('-topmost', True)
                    self.dep_pm.attributes('-fullscreen', True) 
                    self.attributes('-topmost', False)

                    # Load the image file
                    self.image_deppm = Image.open("GUI_images/AttendanceBackgroundConfirm.png")

                    # Create a PhotoImage object from the image
                    self.photo_deppm = ImageTk.PhotoImage(self.image_deppm)

                    # Create a canvas widget with the background image
                    self.canvas_deppm = ctk.CTkCanvas(self.dep_pm, width=self.image_deppm.width, height=self.image_deppm.height)
                    self.canvas_deppm.create_image(0, 0, anchor=ctk.NW, image=self.photo_deppm)
                    self.canvas_deppm.pack()

                    self.webcam_label2 = style.webcam_style_label4(self.dep_pm)
                    self.webcam_label2.place(x=663.4, y=197.4, width=598.8, height=428.7)

                    self.webcam_label2.configure(image=photo)
                    self.webcam_label2.image = photo

                    self.confirm_btn_deppm = style.frame_btn_style(self.dep_pm, "Confirm", self.confirm_func_btn_deppm)
                    self.confirm_btn_deppm.place(x=183.6, y=516.2)

                    self.reset_btn_pm = style.no_btn_style(self.dep_pm, "Reset", self.reset_func_btn_deppm)
                    self.reset_btn_pm.place(x=354.1, y=516.2)
                else:
                    self.attributes('-topmost', False)
                    messagebox.showerror("Error", "No matching image found.")
                    self.attributes('-topmost', True)
                    
    
            else:
                self.attributes('-topmost', False)
                messagebox.showerror("Error", "Unknown Face Detected.")
                self.attributes('-topmost', True)

    def confirm_func_btn_deppm(self):       
        workbook = openpyxl.load_workbook("Csc_Form_48.xlsx")
        worksheet = workbook.get_sheet_by_name("DepPM")
        now = datetime.now()
        row = worksheet.max_row + 1
                    
        worksheet.cell(row=row, column=4, value=now.strftime("%H:%M:%S"))
        worksheet.cell(row=row, column=1, value=self.facerecognized)
        worksheet.cell(row=row, column=2, value=now.strftime("%m/%d/%Y"))
        worksheet.cell(row=row, column=3, value=now.strftime("%a"))
                    
        workbook.save('Csc_Form_48.xlsx')
        self.dep_pm.attributes('-topmost', False)
        messagebox.showinfo("Arrival", "You departed at " + now.strftime("%A") + " , " + now.strftime("%H:%M"))
        self.attributes('-topmost', True)
        self.dep_pm.destroy()

        print("Departure PM just clicked")
    def reset_func_btn_deppm(self):
        self.btn_confirmed = False
        self.webcam_label2.configure(image=None)
        self.webcam_label2.image = None
        self.dep_pm.destroy()
   
    def back_func_btn(self):
        self.capture.release()
        self.destroy()
        print("Back button just clicked")
       